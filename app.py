import requests
from datetime import date, timedelta
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

BASE_URL = f"https://data.xotelo.com/api/rates?currency=USD"

HOTEL_KEYS = {
    "Quality Inn": "g50892-d95503",
    "Comfort Inn": "g50892-d256944",
    "Holiday Inn": "g50892-d631380",
    "Laquinta Macedonia": "g50587-d226031",
    "Comfort Independence": "g50470-d95403"
}

start = 1 #start tomorrow, since api disabled pulling rates for today 
end = 30 #inclusive, run for 30 days

def main():
    today_date = date.today()
    wb, sheet = initialize_workbook()

    intro_string = f'''
    Made with ❤️  by Darshan Thakkar
    This program pulls hotel rates from the Xotelo API for a list of hotels in Richfield, Ohio.
    These rates are gathered from Trip Advisor, which lists rates from OTA sites like Expedia and Booking.com.
    The rates are pulled for a 30-day period starting tomorrow.
    Rates will be saved to file: rates-{today_date}.xlsx
    '''

    print(intro_string)

    for i in range(start, end+1):
        target_date = today_date + timedelta(days=i)
        formatted_date = target_date.strftime("%m/%d")

        date_cell = sheet.cell(row=1+i, column=1, value=formatted_date)
        date_cell.font = Font(bold=True)
        print(f"Pulling rates for date: {formatted_date}")

        for col, (hotel_name, hotel_key) in enumerate(HOTEL_KEYS.items()):
            rate = get_rate(hotel_key, target_date)
            cell = sheet.cell(row=1+i, column=3+col)
            cell.alignment = Alignment(horizontal='center')
            if rate is not None:
                cell.value = rate
            else:
                cell.value = "N/A"

    finalize_sheet(sheet)
    wb.save(f"rates-richfield-{today_date}.xlsx")
    
def initialize_workbook():
    wb = Workbook()
    # wb = load_workbook(f"rates-{today_date}.xlsx")
    sheet = wb.active
    write_header(sheet)
    return wb, sheet

def get_rate(hotel_key, date):
    url = f"{BASE_URL}&hotel_key={hotel_key}&chk_in={date}&chk_out={date + timedelta(days=1)}"

    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
    except requests.exceptions.RequestException as e:
        print(f"Network error for hotel: {hotel_key} on date: {date} - {e}")
        return None
    except json.JSONDecodeError as e:
        print(f"JSON decoding error for hotel: {hotel_key} on date: {date} - {e}")
        return None

    if "error" in data and data["error"] is not None:
        print(f"Error {data["error"]["status_code"]}: {data["error"]["message"]}")
        return None

    if "result" not in data or data["result"] is None:
        print("Error: Missing 'result' in response.")
        return None

    rates = data["result"]["rates"]
    if not rates:
        print(f"No rates found for hotel: {hotel_key} for date: {date}")
        return None
    
    return rates[0]["rate"]

def write_header(sheet):

    sheet.cell( row=1, column=1 , value="Date")
    sheet.cell(row=1, column=2, value = "Availability")

    sheet.cell(row=1, column=3, value = "Quality Inn Richfield")
    sheet.cell(row=1, column=4, value = "Comfort Inn Richfield")
    sheet.cell(row=1, column=5, value = "Holiday Inn Richfield")
    sheet.cell(row=1, column=6, value = "La Quinta Macedonia")
    sheet.cell(row=1, column=7, value = "Comfort Inn Independence")

def finalize_sheet(sheet):

    # bold and center the headers row
    for column in range(1, 8):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    sheet.column_dimensions['A'].width = 9 # date column
    sheet.column_dimensions['B'].width = 14 # availability column
    sheet.column_dimensions['C'].width = 15 
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 16

    # Define a thin border for all sides
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # apply border and font 14 to all cells in the worksheet
    for row in sheet.iter_rows():
        # set row height
        sheet.row_dimensions[row[0].row].height = 21
        
        for cell in row:
            cell.border = thin_border
            cell.font = Font(size=14, bold=cell.font.bold)

    sheet.row_dimensions[1].height = 42

    # set custom margins
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.header = 0.3
    sheet.page_margins.footer = 0.3

if __name__ == "__main__":
    main()